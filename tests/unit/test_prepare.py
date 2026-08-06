"""End-to-end preparation of a corpus, and regression locks on the shipped one."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path

import pytest

from pstparser.config import CorpusSource, load_experiment
from pstparser.data import (
    CorpusError,
    PreparationResult,
    apply_fixes,
    build_target,
    iter_rows,
    load_fixes,
    load_records,
    load_split,
    prepare_corpus,
    read_corpus,
    select,
    sheet_names,
)
from pstparser.pst import serialise_target

#: Digest of the serialised targets produced from the shipped corpus, newline
#: separated. Any change to the taxonomy, to cell normalisation, to the
#: corrections or to the corpus itself moves this value.
SHIPPED_TARGETS_DIGEST = "513e1e40745325afae6130156d5649d2e2cab371ca24034486f76fd64dad5154"

#: Digest of the system message the model is conditioned on.
SYSTEM_PROMPT_DIGEST = "bc2282e04b35c1055a36edc30050286e3ead4bfec159385c09f0462763fd79d7"

SHIPPED_RECORD_COUNT = 975


@pytest.fixture
def prepared(config_dir: Path, tmp_path: Path) -> PreparationResult:
    """Prepare the tiny corpus into a temporary directory."""
    config = load_experiment(
        config_dir / "valid.yaml",
        overrides=[
            f"data.processed_dir={(tmp_path / 'processed').as_posix()}",
            f"data.split.output_dir={(tmp_path / 'splits').as_posix()}",
        ],
        root=config_dir,
    )
    return prepare_corpus(config.data)


def test_every_row_becomes_a_record(prepared: PreparationResult) -> None:
    assert len(prepared.records) == 7
    assert [record.index for record in prepared.records] == list(range(7))


def test_records_carry_the_raw_prompt(prepared: PreparationResult) -> None:
    assert prepared.records[1].prompt == "Fix this bug."


def test_targets_are_parseable(prepared: PreparationResult) -> None:
    for record in prepared.records:
        assert "prompt" in json.loads(record.target)


def test_a_cell_marked_as_disjoint_becomes_two_segments(prepared: PreparationResult) -> None:
    tree = json.loads(prepared.records[2].target)["prompt"]

    assert tree["context"]["data"] == ["class A: pass", "class B: pass"]


def test_a_list_shaped_cell_is_kept_as_the_prompt_writes_it(
    prepared: PreparationResult,
) -> None:
    # The prompt itself contains the brackets and quotes, so the annotation is
    # a faithful span of it. Reading the cell as a literal would replace that
    # span with two words the prompt never held on their own.
    tree = json.loads(prepared.records[6].target)["prompt"]

    assert tree["context"]["data"] == ["['alpha', 'beta']"]


def test_missing_annotation_becomes_an_empty_leaf(prepared: PreparationResult) -> None:
    tree = json.loads(prepared.records[3].target)["prompt"]

    assert tree["main_instruction"] == []
    assert tree["context"]["data"] == ["error: cannot open file"]


def test_under_annotated_row_is_reported(prepared: PreparationResult) -> None:
    # Row 4 is annotated only in part, and is the only row that is.
    assert [issue.index for issue in prepared.quality.issues] == [4]


def test_artefacts_are_written(prepared: PreparationResult) -> None:
    assert prepared.records_path.is_file()
    assert prepared.report_path.is_file()
    assert json.loads(prepared.report_path.read_text(encoding="utf-8"))["total"] == 7


def test_records_round_trip_through_disk(prepared: PreparationResult) -> None:
    assert load_records(prepared.records_path) == prepared.records


def test_partition_is_written_and_reloadable(prepared: PreparationResult) -> None:
    assert load_split(prepared.split_dir) == prepared.split


def test_selection_follows_the_requested_order(prepared: PreparationResult) -> None:
    selected = select(prepared.records, [3, 0])

    assert [record.index for record in selected] == [3, 0]


def second_source(config_dir: Path, tmp_path: Path, prompts: list[str]) -> Path:
    """Write a further annotated worksheet, laid out like the corpus."""
    from openpyxl import Workbook

    config = load_experiment(config_dir / "valid.yaml", root=config_dir)
    headers = [config.data.prompt_column, *config.data.column_mapping.values()]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "extra"
    worksheet.append(headers)
    for prompt in prompts:
        worksheet.append([prompt, prompt, *([None] * 8)])

    destination = tmp_path / "extra.xlsx"
    workbook.save(destination)
    return destination


def test_sources_are_concatenated_and_numbered_across(config_dir: Path, tmp_path: Path) -> None:
    extra = second_source(config_dir, tmp_path, ["Rename the variable.", "Delete the file."])
    config = load_experiment(
        config_dir / "valid.yaml",
        overrides=[
            f"data.processed_dir={(tmp_path / 'processed').as_posix()}",
            f"data.split.output_dir={(tmp_path / 'splits').as_posix()}",
        ],
        root=config_dir,
    )
    sources = [*config.data.sources, CorpusSource(path=extra, sheet="extra")]

    result = prepare_corpus(config.data.model_copy(update={"sources": sources}))

    assert len(result.records) == 9
    assert [record.index for record in result.records] == list(range(9))
    assert result.records[7].prompt == "Rename the variable."
    assert result.records[8].prompt == "Delete the file."


def test_a_source_can_carry_no_corrections(config_dir: Path, tmp_path: Path) -> None:
    extra = second_source(config_dir, tmp_path, [f"Handle case {n}." for n in range(8)])
    config = load_experiment(
        config_dir / "valid.yaml",
        overrides=[
            f"data.processed_dir={(tmp_path / 'processed').as_posix()}",
            f"data.split.output_dir={(tmp_path / 'splits').as_posix()}",
        ],
        root=config_dir,
    )
    only_extra = [CorpusSource(path=extra, sheet="extra")]

    result = prepare_corpus(config.data.model_copy(update={"sources": only_extra}))

    assert result.fixes == []
    assert result.quality.passed


def annotated_source(config_dir: Path, tmp_path: Path) -> tuple[Path, CorpusSource]:
    """Copy the text-annotated fixture next to the other temporary files."""
    destination = tmp_path / "annotated.yaml"
    destination.write_text(
        Path("tests/assets/tiny_annotations.yaml").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return destination, CorpusSource(path=destination, format="yaml")


def test_a_text_source_prepares_on_its_own(config_dir: Path, tmp_path: Path) -> None:
    _, source = annotated_source(config_dir, tmp_path)
    config = load_experiment(
        config_dir / "valid.yaml",
        overrides=[
            f"data.processed_dir={(tmp_path / 'processed').as_posix()}",
            f"data.split.output_dir={(tmp_path / 'splits').as_posix()}",
        ],
        root=config_dir,
    )

    result = prepare_corpus(config.data.model_copy(update={"sources": [source]}))

    assert [record.index for record in result.records] == [0, 1, 2]
    # The fixture is written to satisfy the extraction contract, so the same
    # check that guards the delivered corpus passes on it unchanged.
    assert result.quality.passed
    assert result.fixes == []


def test_a_text_source_concatenates_after_a_worksheet(config_dir: Path, tmp_path: Path) -> None:
    _, source = annotated_source(config_dir, tmp_path)
    config = load_experiment(
        config_dir / "valid.yaml",
        overrides=[
            f"data.processed_dir={(tmp_path / 'processed').as_posix()}",
            f"data.split.output_dir={(tmp_path / 'splits').as_posix()}",
        ],
        root=config_dir,
    )
    sources = [*config.data.sources, source]

    result = prepare_corpus(config.data.model_copy(update={"sources": sources}))

    assert len(result.records) == 10
    assert [record.index for record in result.records] == list(range(10))
    assert result.records[7].prompt.startswith("You are a senior Go reviewer.")
    # Which file a record came from, which its corpus-wide index no longer says.
    assert [count for _, count in result.counts] == [7, 3]
    assert Path(result.counts[1][0]).name == "annotated.yaml"


def test_a_text_record_serialises_in_taxonomy_order(config_dir: Path, tmp_path: Path) -> None:
    _, source = annotated_source(config_dir, tmp_path)
    config = load_experiment(
        config_dir / "valid.yaml",
        overrides=[
            f"data.processed_dir={(tmp_path / 'processed').as_posix()}",
            f"data.split.output_dir={(tmp_path / 'splits').as_posix()}",
        ],
        root=config_dir,
    )

    result = prepare_corpus(config.data.model_copy(update={"sources": [source]}))
    target = json.loads(result.records[0].target)

    assert list(target["prompt"]) == ["main_instruction", "context", "examples", "reasoning"]
    assert target["prompt"]["reasoning"]["influence"] == ["Think step by step."]


def test_the_paradigm_travels_with_the_record(config_dir: Path, tmp_path: Path) -> None:
    _, source = annotated_source(config_dir, tmp_path)
    config = load_experiment(
        config_dir / "valid.yaml",
        overrides=[
            f"data.processed_dir={(tmp_path / 'processed').as_posix()}",
            f"data.split.output_dir={(tmp_path / 'splits').as_posix()}",
        ],
        root=config_dir,
    )

    result = prepare_corpus(config.data.model_copy(update={"sources": [source]}))
    lines = result.records_path.read_text(encoding="utf-8").splitlines()

    assert [record.paradigm for record in result.records] == [
        "zero_shot_cot",
        None,
        "few_shot_cot",
    ]
    assert json.loads(lines[0])["paradigm"] == "zero_shot_cot"
    # A record whose source records no paradigm serialises exactly as it did
    # before the field existed, rather than gaining a null.
    assert "paradigm" not in json.loads(lines[1])
    assert load_records(result.records_path) == result.records


def test_a_worksheet_record_carries_no_paradigm(prepared: PreparationResult) -> None:
    lines = prepared.records_path.read_text(encoding="utf-8").splitlines()

    assert all(record.paradigm is None for record in prepared.records)
    assert all("paradigm" not in json.loads(line) for line in lines)


def test_a_spreadsheet_source_without_a_worksheet_is_rejected(
    config_dir: Path, tmp_path: Path
) -> None:
    config = load_experiment(config_dir / "valid.yaml", root=config_dir)
    # The schema forbids it; model_copy does not revalidate, which is what makes
    # the guard in the reader reachable rather than dead.
    source = config.data.sources[0].model_copy(update={"sheet": None})

    with pytest.raises(CorpusError, match="must name a worksheet"):
        prepare_corpus(config.data.model_copy(update={"sources": [source]}))


def test_incomplete_column_mapping_is_rejected(config_dir: Path, tmp_path: Path) -> None:
    config = load_experiment(config_dir / "valid.yaml", root=config_dir)
    mapping = dict(config.data.column_mapping)
    del mapping["context.role"]
    broken = config.data.model_copy(update={"column_mapping": mapping})

    with pytest.raises(CorpusError, match="does not cover"):
        prepare_corpus(broken)


def test_failing_integrity_check_can_abort(config_dir: Path, tmp_path: Path) -> None:
    config = load_experiment(
        config_dir / "valid.yaml",
        overrides=[
            "data.quality.fail_on_issues=true",
            f"data.processed_dir={(tmp_path / 'processed').as_posix()}",
        ],
        root=config_dir,
    )

    with pytest.raises(CorpusError, match="integrity check reported"):
        prepare_corpus(config.data)


def test_unknown_worksheet_is_rejected(tiny_corpus: Path) -> None:
    with pytest.raises(CorpusError, match="worksheet 'absent' not found"):
        read_corpus(tiny_corpus, "absent", required_columns=[])


def test_missing_column_is_rejected(tiny_corpus: Path) -> None:
    with pytest.raises(CorpusError, match="missing columns"):
        read_corpus(tiny_corpus, "corpus", required_columns=["NO SUCH COLUMN"])


def test_worksheets_are_listed(tiny_corpus: Path) -> None:
    assert sheet_names(tiny_corpus) == ["corpus"]


def test_system_prompt_is_unchanged() -> None:
    digest = hashlib.sha256(Path("prompts/system_pst.md").read_bytes()).hexdigest()

    assert digest == SYSTEM_PROMPT_DIGEST


def test_shipped_corpus_produces_stable_targets() -> None:
    config = load_experiment("configs/experiments/baseline.yaml")
    frame = read_corpus(
        config.data.sources[0].path,
        config.data.sources[0].sheet,
        required_columns=[config.data.prompt_column, *config.data.column_mapping.values()],
    )
    fixes = config.data.sources[0].fixes
    assert fixes is not None
    frame = apply_fixes(frame, load_fixes(fixes))

    targets = [
        serialise_target(build_target(row, config.data.column_mapping)) for row in iter_rows(frame)
    ]
    digest = hashlib.sha256("\n".join(targets).encode("utf-8")).hexdigest()

    assert len(targets) == SHIPPED_RECORD_COUNT
    assert digest == SHIPPED_TARGETS_DIGEST
