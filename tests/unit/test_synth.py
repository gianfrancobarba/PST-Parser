"""Corpus expansion, exercised against a stand-in provider."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

from pstparser.config import SynthConfig
from pstparser.config.loader import load_experiment
from pstparser.pst import LEAF_PATHS
from pstparser.synth import (
    ProviderError,
    SyntheticPrompt,
    generate_prompts,
    load_seeds,
    provider_from_env,
    run_synthesis,
    write_annotation_sheet,
)

SEEDS = {
    "zero_shot_cot": ["Think through the migration before fixing it."],
    "tree_of_thoughts": ["Generate three options and choose one."],
}


class ScriptedProvider:
    """Returns queued answers in order, recording what it was asked."""

    def __init__(self, answers: list[str]) -> None:
        """Queue the answers the provider will hand out."""
        self.answers = list(answers)
        self.calls: list[tuple[str, str, float]] = []

    def complete(self, system: str, user: str, temperature: float) -> str:
        """Pop the next queued answer."""
        self.calls.append((system, user, temperature))
        return self.answers.pop(0) if self.answers else ""


class FailingProvider:
    """Always refuses."""

    def complete(self, system: str, user: str, temperature: float) -> str:
        """Raise, as a provider does when the service is unreachable."""
        raise ProviderError("unreachable")


def test_seeds_are_read_and_grouped() -> None:
    seeds = load_seeds("data/seeds/paradigms.yaml")

    assert set(seeds) == {"zero_shot_cot", "few_shot_cot", "tree_of_thoughts"}
    assert all(len(entries) == 10 for entries in seeds.values())
    assert all(text.strip() for entries in seeds.values() for text in entries)


def test_missing_seed_file_is_rejected(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError, match="seed file not found"):
        load_seeds(tmp_path / "absent.yaml")


def test_seed_file_without_entries_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "seeds.yaml"
    path.write_text(yaml.safe_dump({"zero_shot_cot": []}), encoding="utf-8")

    with pytest.raises(ValueError, match="has no seeds"):
        load_seeds(path)


def test_one_prompt_is_requested_per_paradigm_and_slot() -> None:
    provider = ScriptedProvider(["a", "b", "c", "d"])

    prompts, rejected = generate_prompts(SEEDS, provider, SynthConfig(per_paradigm=2))

    assert len(prompts) == 4
    assert rejected == 0
    assert len(provider.calls) == 4


def test_generated_prompts_carry_their_paradigm() -> None:
    provider = ScriptedProvider(["one", "two"])

    prompts, _ = generate_prompts(SEEDS, provider, SynthConfig(per_paradigm=1))

    assert [prompt.paradigm for prompt in prompts] == ["zero_shot_cot", "tree_of_thoughts"]


def test_request_shows_the_seeds_of_that_paradigm() -> None:
    provider = ScriptedProvider(["one", "two"])

    generate_prompts(SEEDS, provider, SynthConfig(per_paradigm=1))

    first_request = provider.calls[0][1]
    assert "zero shot cot" in first_request
    assert SEEDS["zero_shot_cot"][0] in first_request
    assert SEEDS["tree_of_thoughts"][0] not in first_request


def test_temperature_is_passed_through() -> None:
    provider = ScriptedProvider(["one", "two"])

    generate_prompts(SEEDS, provider, SynthConfig(per_paradigm=1, temperature=0.4))

    assert all(call[2] == pytest.approx(0.4) for call in provider.calls)


def test_duplicate_completions_are_discarded() -> None:
    provider = ScriptedProvider(["same", "SAME", "other", "x"])

    prompts, rejected = generate_prompts(
        {"zero_shot_cot": SEEDS["zero_shot_cot"]}, provider, SynthConfig(per_paradigm=4)
    )

    assert [prompt.text for prompt in prompts] == ["same", "other", "x"]
    assert rejected == 1


def test_completion_repeating_a_seed_is_discarded() -> None:
    provider = ScriptedProvider([SEEDS["zero_shot_cot"][0], "fresh"])

    prompts, rejected = generate_prompts(
        {"zero_shot_cot": SEEDS["zero_shot_cot"]}, provider, SynthConfig(per_paradigm=2)
    )

    assert [prompt.text for prompt in prompts] == ["fresh"]
    assert rejected == 1


def test_empty_completion_is_discarded() -> None:
    provider = ScriptedProvider(["", "   ", "kept"])

    prompts, rejected = generate_prompts(
        {"zero_shot_cot": SEEDS["zero_shot_cot"]}, provider, SynthConfig(per_paradigm=3)
    )

    assert [prompt.text for prompt in prompts] == ["kept"]
    assert rejected == 2


def test_provider_failures_do_not_abort_the_run() -> None:
    prompts, rejected = generate_prompts(SEEDS, FailingProvider(), SynthConfig(per_paradigm=2))

    assert prompts == []
    assert rejected == 4


def test_annotation_sheet_matches_the_corpus_layout(tmp_path: Path) -> None:
    config = load_experiment("configs/experiments/baseline.yaml")
    prompts = [
        SyntheticPrompt(paradigm="zero_shot_cot", text="first"),
        SyntheticPrompt(paradigm="tree_of_thoughts", text="second"),
    ]

    path = write_annotation_sheet(
        prompts,
        tmp_path / "nested" / "sheet.xlsx",
        column_mapping=config.data.column_mapping,
        prompt_column=config.data.prompt_column,
        sheet_name="synthetic",
    )

    worksheet = load_workbook(path)["synthetic"]
    headers = [cell.value for cell in worksheet[1]]

    assert headers[0] == "PROMPT"
    assert len(headers) == len(LEAF_PATHS) + 1
    assert headers[1:] == [config.data.column_mapping[path] for path in LEAF_PATHS]
    assert [row[0].value for row in worksheet.iter_rows(min_row=2)] == ["first", "second"]
    # The annotation columns are left for a human to fill in.
    assert all(cell.value is None for cell in worksheet[2][1:])


def test_missing_credential_is_reported() -> None:
    with pytest.raises(ProviderError, match="ABSENT_VARIABLE is not set"):
        provider_from_env("https://example.invalid/v1", "some/model", "ABSENT_VARIABLE")


def test_run_writes_sheet_summary_and_manifest(tmp_path: Path) -> None:
    config = load_experiment(
        "configs/experiments/baseline.yaml",
        overrides=[
            f"synth.output_dir={(tmp_path / 'synthetic').as_posix()}",
            "synth.per_paradigm=1",
        ],
    )
    provider = ScriptedProvider([f"generated {n}" for n in range(10)])

    result = run_synthesis(config, provider=provider)

    assert result.requested == 3
    assert len(result.prompts) == 3
    assert result.per_paradigm() == {
        "zero_shot_cot": 1,
        "few_shot_cot": 1,
        "tree_of_thoughts": 1,
    }
    assert result.sheet_path.is_file()

    summary = json.loads((tmp_path / "synthetic" / "synthesis_summary.json").read_text("utf-8"))
    assert summary["accepted"] == 3

    manifest = json.loads((tmp_path / "synthetic" / "run_manifest.json").read_text("utf-8"))
    assert manifest["stage"] == "synth"
    assert manifest["inputs"]["seeds"]
    assert manifest["synthesis"]["accepted"] == 3
